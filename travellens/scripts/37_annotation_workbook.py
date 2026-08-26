"""Build an Excel workbook to label the gold set by hand.

  python scripts/37_annotation_workbook.py
  python scripts/37_annotation_workbook.py --annotator 2

Same job as 35_annotate.py and 36_annotation_sheet.py; this is the version for
someone who would rather work in Excel than a terminal or a raw CSV. It adds
what a CSV cannot carry:

  * dropdowns on every answer cell, so N/P/X can be picked rather than typed
    and a typo is not possible
  * wrapped text and set column widths, so a long review is readable instead
    of running off the screen
  * a frozen header, so the column names stay visible at row 180
  * an instructions sheet with worked examples from this very file

It deliberately does NOT include sample_reason -- that column records which
answer the pipeline expects, and an annotator who can see it is no longer an
independent one.

Read the filled workbook back with:
  python scripts/37_annotation_workbook.py --import-from <file>
"""
import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import pandas as pd  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

from travellens import config as C  # noqa: E402

# The four the focused set was sampled to measure.
ASPECTS = ["roads_access", "cleanliness", "facilities", "safety"]
# The other three. Not part of the sampling design -- price, crowding and
# scenery already score 0.98, 0.90 and 0.91 with the plain word list, so there
# was little to check. But a piece like "there were no much crowd when we
# reach there" is plainly ABOUT crowding, and having nowhere to say so makes
# the task feel wrong even when a blank is the correct answer. --all-aspects
# adds them, and anything recorded in them is a bonus rather than part of the
# headline measurement.
EXTRA = ["price_value", "crowd", "scenery"]
NICE = {"roads_access": "Roads & access", "cleanliness": "Cleanliness",
        "facilities": "Facilities", "safety": "Safety",
        "price_value": "Price & value", "crowd": "Crowding & noise",
        "scenery": "Scenery"}
VALID = {"N", "P", "X"}

COVERS = {
    "roads_access": "roads, parking, buses, the walk, signage, finding it",
    "cleanliness": "litter, plastic, smells, upkeep",
    "facilities": "toilets, food, seating, shelter, bins, guides",
    "safety": "slippery ground, deep water, wildlife, warnings",
    "price_value": "entrance fees, parking charges, value for money",
    "crowd": "busy, queues, noise -- and also peaceful or quiet",
    "scenery": "the view, the landscape, how beautiful it is",
}

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3B36")
ANSWER_FILL = PatternFill("solid", fgColor="FFF7D6")   # the cells to fill in
THIN = Side(style="thin", color="C8D2CF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def goldset_path(annotator, full=False):
    stem = "goldset_annotator" if full else "goldset_focused_annotator"
    return C.REPORTS / "{}{}.csv".format(stem, annotator)


def _instructions(ws, examples, aspects):
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 104
    rows = [
        ("h1", "How to label these 200 rows"),
        ("p", ""),
        ("p", "You are reading one PIECE of a review and saying whether the "
              "visitor is complaining about any of {} things.".format(
                  len(aspects))),
        ("p", ""),
        ("h2", "The {} columns".format(len(aspects))),
    ] + [
        ("p", "{:<18} {}".format(NICE[a], COVERS[a])) for a in aspects
    ] + [
        ("p", ""),
        ("h2", "What to put in a cell"),
        ("p", "N   the visitor is COMPLAINING about it"),
        ("p", "P   the visitor is PRAISING it"),
        ("p", "X   mentioned, but no opinion either way"),
        ("p", "    leave it EMPTY if that thing is not mentioned at all"),
        ("p", ""),
        ("p", "Each cell has a dropdown, so you can pick instead of typing."),
        ("p", "MOST CELLS STAY EMPTY. A typical piece touches one thing, often none."),
        ("p", ""),
        ("h2", "The one rule people get wrong"),
        ("p", "Judge the PIECE column only. The whole review is shown beside it "
              "so you can tell what the piece means -- it is NOT what you are "
              "judging. A review can complain about the road in a sentence that "
              "is not the piece in front of you."),
        ("p", ""),
        ("h2", "Worked examples, from this file"),
    ]
    for kind, text in rows:
        ws.append([None, text])
        cell = ws.cell(row=ws.max_row, column=2)
        cell.font = Font(name=FONT, size=14 if kind == "h1" else 11,
                         bold=kind in ("h1", "h2"),
                         color="1F3B36" if kind != "p" else "000000")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.append([])
    ws.append([None, "PIECE", "Answer", "Why"])
    for col in (2, 3, 4):
        c = ws.cell(row=ws.max_row, column=col)
        c.font = Font(name=FONT, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
    for piece, answer, why in examples:
        ws.append([None, piece, answer, why])
        for col in (2, 3, 4):
            c = ws.cell(row=ws.max_row, column=col)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 46
    ws.sheet_view.showGridLines = False


def export(annotator=1, full=False, all_aspects=False):
    src = goldset_path(annotator, full)
    df = pd.read_csv(src)
    aspects = [a for a in ASPECTS if a in df.columns] or list(ASPECTS)
    if all_aspects:
        aspects = aspects + [a for a in EXTRA if a not in aspects]

    wb = Workbook()
    _instructions(wb.active, [
        ("We saw a few monkeys, a deer and a few birds",
         "all four empty",
         "Describing wildlife they saw. Not a complaint about anything. "
         "(A wildlife WARNING would be safety; a sighting is not.)"),
        ("On the way you will get a Public Toilet which is not very clean "
         "& hygienic with no water tap in it.",
         "Cleanliness = N, Facilities = N",
         "Two complaints in one piece: the toilet is dirty, and it lacks a tap."),
        ("When you climb up the stairs.",
         "all four empty",
         "Half a sentence with no opinion in it."),
        ("Avoid littering.",
         "usually empty, or Cleanliness = X",
         "The visitor is asking others to behave, not reporting that the "
         "place is dirty. Judgement calls like this are why a person does "
         "this and not a program."),
    ], aspects)
    wb.active.title = "Read me first"

    ws = wb.create_sheet("Label these")
    headers = ["row", "segment_id", "Destination",
               "PIECE  <-- judge THIS", "The whole review (context only)"]
    headers += [NICE[a] for a in aspects] + ["notes (optional)"]
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = HEAD_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BORDER
    ws.row_dimensions[1].height = 34

    first_aspect = 6
    for r in df.to_dict("records"):
        ws.append([r["row"], r["segment_id"], r.get("destination", ""),
                   str(r.get("segment", "")),
                   str(r.get("full_review", "") or "")]
                  + ["" for _ in aspects] + [""])

    widths = {1: 6, 2: 20, 3: 24, 4: 52, 5: 64}
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w
    for n in range(len(aspects)):
        ws.column_dimensions[get_column_letter(first_aspect + n)].width = 13
    ws.column_dimensions[get_column_letter(first_aspect + len(aspects))].width = 26

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDER
        for n in range(len(aspects)):
            a = row[first_aspect - 1 + n]
            a.fill = ANSWER_FILL
            a.alignment = Alignment(horizontal="center", vertical="center")
            a.font = Font(name=FONT, size=11, bold=True)

    # Dropdowns: pick, do not type. Removes the whole class of typo the CSV
    # importer had to reject.
    dv = DataValidation(type="list", formula1='"N,P,X"', allow_blank=True,
                        showDropDown=False)
    dv.error = "Use N (complaining), P (praising), X (mentioned, no opinion), or leave empty."
    dv.errorTitle = "One of N, P, X"
    dv.prompt = "N = complaining, P = praising, X = mentioned only. Empty = not mentioned."
    dv.promptTitle = "Your verdict"
    ws.add_data_validation(dv)
    last = ws.max_row
    for n in range(len(aspects)):
        col = get_column_letter(first_aspect + n)
        dv.add("{}2:{}{}".format(col, col, last))

    # segment_id is how the import finds the row again; it is not for reading.
    ws.column_dimensions["B"].hidden = True
    ws.freeze_panes = "F2"
    ws.sheet_view.showGridLines = False

    dest = C.REPORTS / "LABEL_THESE_annotator{}{}.xlsx".format(
        annotator, "_all7" if all_aspects else "")
    wb.save(dest)
    return dest, len(df)


def read_back(filled, annotator=1, full=False):
    src = Path(filled)
    if not src.exists():
        sys.exit("no such file: {}".format(src))
    wb = load_workbook(src, data_only=True)
    ws = wb["Label these"] if "Label these" in wb.sheetnames else wb.active

    header = [c.value for c in ws[1]]
    try:
        id_col = header.index("segment_id") + 1
    except ValueError:
        sys.exit("that workbook has no segment_id column -- is it the right file?")
    label_cols = {}
    for a in ASPECTS + EXTRA:
        if NICE[a] in header:
            label_cols[a] = header.index(NICE[a]) + 1
    if not label_cols:
        sys.exit("no answer columns found in that workbook")
    notes_col = (header.index("notes (optional)") + 1
                 if "notes (optional)" in header else None)

    answers, problems = {}, []
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(row=r, column=id_col).value
        if sid is None:
            continue
        vals, bad = {}, False
        for a, col in label_cols.items():
            raw = ws.cell(row=r, column=col).value
            v = "" if raw is None else str(raw).strip().upper()
            if v in ("", "-"):
                vals[a] = ""
            elif v in VALID:
                vals[a] = v
            else:
                problems.append("row {}: {} = '{}'".format(
                    ws.cell(row=r, column=1).value, NICE[a], raw))
                bad = True
        if not bad:
            note = ws.cell(row=r, column=notes_col).value if notes_col else None
            answers[str(sid)] = (vals, "" if note is None else str(note))

    if problems:
        print("\n  NOT IMPORTED -- {} cell(s) I could not read:".format(len(problems)))
        for p in problems[:20]:
            print("    {}".format(p))
        print("\n  Fix those and run this again. Nothing was written.")
        return None

    gold = pd.read_csv(goldset_path(annotator, full))
    # A column the sheet carries but the gold set does not is added rather
    # than dropped: a verdict somebody actually made must not be thrown away
    # because the original sampling did not ask for it.
    for a in list(label_cols) + ["checked", "notes"]:
        if a not in gold.columns:
            gold[a] = ""
        gold[a] = gold[a].astype("object")

    filled_rows = blank_rows = 0
    for i, r in gold.iterrows():
        sid = str(r["segment_id"])
        if sid not in answers:
            continue
        vals, note = answers[sid]
        # An untouched workbook must import nothing. 200 empty cells are not
        # 200 verdicts of "nothing here" -- they are a file nobody read.
        if not any(vals.values()) and not note.strip():
            blank_rows += 1
            continue
        for a in label_cols:
            gold.at[i, a] = vals[a]
        gold.at[i, "notes"] = note
        gold.at[i, "checked"] = "x"
        filled_rows += 1

    dest = goldset_path(annotator, full)
    gold.to_csv(dest, index=False, encoding="utf-8")
    print("\n  imported {} labelled rows -> {}".format(filled_rows, dest.name))
    if blank_rows:
        print("  {} rows were entirely empty and were SKIPPED.".format(blank_rows))
        print("  If a blank row really is your verdict, put a note on it.")
    print("\n  next: python scripts/05_check_goldset.py")
    return dest


def main():
    ap = argparse.ArgumentParser(description="Excel workbook for annotation.")
    ap.add_argument("--annotator", type=int, default=1, choices=(1, 2))
    ap.add_argument("--import-from", dest="import_from")
    ap.add_argument("--full", action="store_true")
    ap.add_argument("--all-aspects", dest="all_aspects", action="store_true",
                    help="add price, crowding and scenery columns")
    args = ap.parse_args()

    print("\nLostinSriLanka -- annotation workbook\n" + "=" * 60)
    if args.import_from:
        read_back(args.import_from, args.annotator, args.full)
        return
    dest, n = export(args.annotator, args.full, args.all_aspects)
    print("\n  wrote {}".format(dest))
    print("  {} rows, dropdowns on every answer cell".format(n))
    print("\n  Open it, read the first tab, fill in the second.")
    print("  Then: python scripts/37_annotation_workbook.py --import-from \"{}\""
          .format(dest.name))


if __name__ == "__main__":
    main()
