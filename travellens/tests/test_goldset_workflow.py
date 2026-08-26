"""
Guard tests: the path from labelling to a reported kappa must not break.

Open problem #1 is the highest-value outstanding task in this repository, and
the tooling around it had two faults that would only surface AFTER somebody
spent an hour labelling: the checker read a different file than the annotator
writes, and it assumed seven aspect columns when the focused set has four.
Both are the kind of thing that wastes a person's afternoon, so both are
pinned here.

Run:  python -m pytest tests/test_goldset_workflow.py -q
"""
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import pandas as pd  # noqa: E402

from travellens import gold_check  # noqa: E402

FOCUSED = ROOT / "reports" / "goldset_focused_annotator1.csv"
FULL = ROOT / "reports" / "goldset_annotator1.csv"


def test_the_checker_reads_the_set_the_annotator_writes():
    """35_annotate.py fills the FOCUSED set by default; so must the checker."""
    src = (ROOT / "src" / "travellens" / "gold_check.py").read_text(encoding="utf-8")
    assert "goldset_focused_annotator" in src
    ann = (ROOT / "scripts" / "35_annotate.py").read_text(encoding="utf-8")
    assert "goldset_focused_annotator" in ann


def test_validate_handles_the_four_aspect_focused_set():
    """It used to raise KeyError on the set the README asks people to label."""
    v = gold_check.validate(FOCUSED)
    assert v["rows"] == 200
    assert set(v["aspects"]) == {"roads_access", "cleanliness",
                                 "facilities", "safety"}


def test_validate_still_handles_the_seven_aspect_set():
    v = gold_check.validate(FULL)
    assert v["rows"] == 600
    assert len(v["aspects"]) == 7


def test_agreement_only_scores_aspects_both_sheets_have(tmp_path):
    """Scoring absent columns produced kappa 1.0 -- perfect agreement about
    nothing, which silently inflated the mean."""
    a = pd.read_csv(FOCUSED).head(40).copy()
    b = a.copy()
    a["safety"] = ["N", "P", "X", ""] * 10
    b["safety"] = ["N", "P", "X", ""] * 10
    b.loc[0, "safety"] = "P"          # one real disagreement
    pa, pb = tmp_path / "a.csv", tmp_path / "b.csv"
    a.to_csv(pa, index=False)
    b.to_csv(pb, index=False)

    out = gold_check.agreement(pa, pb)
    assert set(out["per_aspect"]) == {"roads_access", "cleanliness",
                                      "facilities", "safety"}
    assert "price_value" not in out["per_aspect"]
    assert out["per_aspect"]["safety"]["kappa"] < 1.0


def test_kappa_rewards_agreeing_that_an_aspect_is_absent():
    """Most cells are blank. If blank were dropped rather than treated as its
    own category, the measurement would rest on a small fraction of the set."""
    blank = pd.Series([None] * 30)
    assert gold_check.cohens_kappa(blank, blank) == 1.0


def test_a_disagreement_is_not_reported_as_agreement():
    a = pd.Series(["N"] * 15 + ["P"] * 15)
    b = pd.Series(["P"] * 15 + ["N"] * 15)
    assert gold_check.cohens_kappa(a, b) < 0


def test_the_gold_set_is_still_unlabelled():
    """A tripwire, not a preference.

    These labels must come from a person who did not build the pipeline. If
    this test ever fails because rows were filled in programmatically, the
    circularity that open problem #1 exists to remove has been reintroduced
    and every accuracy figure derived from it is void.
    """
    v = gold_check.validate(FOCUSED)
    if v["rows_checked"]:
        assert v["rows_checked"] == v["rows"], (
            "the gold set is partly labelled -- finish it by hand "
            "(python scripts/35_annotate.py), and never fill it in with code")


def test_the_annotator_survives_a_windows_console():
    """It died four rows in, on a review containing emoji.

    Windows consoles default to cp1252 and the corpus is full of flags and
    leaves. The crash lost the session, and a tool that dies partway through
    is a tool nobody finishes. stdout is reconfigured to UTF-8 with
    errors="replace", and every line that prints review text goes through
    say(), which cannot raise.
    """
    src = (ROOT / "scripts" / "35_annotate.py").read_text(encoding="utf-8")
    assert "_make_console_utf8" in src
    assert 'errors="replace"' in src

    # No bare print() may touch review text: those lines carry the emoji.
    body = src.split("def show(")[1].split("def main(")[0]
    assert "print(" not in body, \
        "show() prints review text with a bare print(); use say()"


def test_say_cannot_raise_on_unencodable_text():
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "annotate", ROOT / "scripts" / "35_annotate.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    mod.say("Avoid littering.\U0001F343\U0001F1F1\U0001F1F0 …")


# --------------------------------------------------------------------------
# The spreadsheet route (scripts/36_annotation_sheet.py)
# --------------------------------------------------------------------------
def _sheet_module():
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "sheet", ROOT / "scripts" / "36_annotation_sheet.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def test_the_sheet_never_carries_the_pipelines_guess():
    """sample_reason says which answer is expected. An annotator who can see
    it is not independent, and independence is the whole exercise."""
    sheet = ROOT / "reports" / "TO_LABEL_annotator1.csv"
    if not sheet.exists():
        _sheet_module().export(1)
    cols = pd.read_csv(sheet, encoding="utf-8-sig", nrows=1).columns
    assert "sample_reason" not in cols
    assert not any("sample" in c.lower() for c in cols)


def test_the_sheet_is_readable_by_excel():
    """Without the BOM, Excel reads it in the local codepage and every
    Sinhala name and emoji becomes mojibake."""
    sheet = ROOT / "reports" / "TO_LABEL_annotator1.csv"
    if not sheet.exists():
        _sheet_module().export(1)
    assert sheet.read_bytes()[:3] == b"\xef\xbb\xbf", "no UTF-8 BOM"


def _gold_copy(tmp_path):
    """A throwaway copy of the gold set.

    Every import test used to write into the REAL file and assert it stayed
    empty. That passed only while it WAS empty; the day it held 200 hand-made
    labels, the same tests failed AND one of them was two lines from erasing
    the lot. Tests never touch the real gold set now.
    """
    src = ROOT / "reports" / "goldset_focused_annotator1.csv"
    dest = tmp_path / "gold.csv"
    dest.write_text(src.read_text(encoding="utf-8"), encoding="utf-8")
    return dest


def _checked(path):
    d = pd.read_csv(path)
    return int((d["checked"].astype(str).str.strip().str.lower() == "x").sum())


def test_an_untouched_sheet_imports_nothing(tmp_path):
    """An empty spreadsheet must not become 200 confident 'nothing here'
    verdicts. That would fabricate a gold set out of a file nobody read."""
    mod = _sheet_module()
    gold = _gold_copy(tmp_path)
    blank = pd.read_csv(ROOT / "reports" / "TO_LABEL_annotator1.csv",
                        encoding="utf-8-sig")
    for a in ["roads_access", "cleanliness", "facilities", "safety"]:
        blank[a] = ""
    path = tmp_path / "untouched.csv"
    blank.to_csv(path, index=False, encoding="utf-8-sig")

    before = _checked(gold)
    mod.read_back(path, 1, gold_path=gold)
    assert _checked(gold) == before, "an unread sheet was imported as verdicts"


def test_an_unreadable_cell_blocks_the_whole_import(tmp_path):
    """Half-importing leaves the gold set in a state nobody chose."""
    mod = _sheet_module()
    gold = _gold_copy(tmp_path)
    sheet = pd.read_csv(ROOT / "reports" / "TO_LABEL_annotator1.csv",
                        encoding="utf-8-sig")
    sheet["safety"] = sheet["safety"].astype("object")
    sheet.loc[0, "safety"] = "N"
    sheet.loc[1, "safety"] = "probably?"
    path = tmp_path / "bad.csv"
    sheet.to_csv(path, index=False, encoding="utf-8-sig")

    before = _checked(gold)
    assert mod.read_back(path, 1, gold_path=gold) is None, "a bad sheet was accepted"
    assert _checked(gold) == before, "a refused import still wrote rows"


# --------------------------------------------------------------------------
# The Excel route (scripts/37_annotation_workbook.py)
# --------------------------------------------------------------------------
def _wb_module():
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "wbmod", ROOT / "scripts" / "37_annotation_workbook.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def test_the_workbook_constrains_answers_to_npx(tmp_path):
    """A dropdown removes the whole class of typo the CSV importer rejects."""
    from openpyxl import load_workbook
    path, _ = _wb_module().export(1, gold_path=_gold_copy(tmp_path))
    dvs = load_workbook(path)["Label these"].data_validations.dataValidation
    assert dvs and dvs[0].formula1 == '"N,P,X"'
    assert dvs[0].allow_blank, "blank must stay allowed -- most cells are blank"


def test_the_workbook_does_not_show_the_expected_answer(tmp_path):
    """sample_reason names the aspect the pipeline thinks the row is about."""
    from openpyxl import load_workbook
    path, _ = _wb_module().export(1, gold_path=_gold_copy(tmp_path))
    wb = load_workbook(path)
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(max_row=3):
            for c in row:
                text = str(c.value or "").lower()
                assert "representative:" not in text
                assert "disagreement:" not in text


def test_the_workbook_names_which_column_to_judge(tmp_path):
    """The commonest mistake is labelling the whole review."""
    from openpyxl import load_workbook
    path, _ = _wb_module().export(1, gold_path=_gold_copy(tmp_path))
    header = [c.value for c in load_workbook(path)["Label these"][1]]
    piece = [h for h in header if h and "PIECE" in h]
    assert piece and "judge" in piece[0].lower()
    assert any(h and "whole review" in h for h in header)


def test_an_untouched_workbook_imports_nothing(tmp_path):
    """200 empty cells are a file nobody read, not 200 verdicts."""
    mod = _wb_module()
    gold = _gold_copy(tmp_path)
    path, _ = mod.export(1, gold_path=gold)
    before = _checked(gold)
    mod.read_back(path, 1, gold_path=gold)
    assert _checked(gold) == before


def test_a_blank_row_before_a_labelled_one_is_a_verdict(tmp_path):
    """40 of annotator 1's 200 rows were blank, scattered throughout, with
    labels running to row 200. Those are judgements of 'nothing applies'.
    Dropping them would keep only rows where the annotator found something,
    biasing the set towards positives."""
    from openpyxl import load_workbook
    mod = _wb_module()
    gold = _gold_copy(tmp_path)
    # start from an unlabelled copy so the high-water mark is ours
    d = pd.read_csv(gold)
    for col in ["roads_access", "cleanliness", "facilities", "safety",
                "checked", "notes"]:
        if col in d.columns:
            d[col] = ""
    d.to_csv(gold, index=False, encoding="utf-8")

    path, _ = mod.export(1, gold_path=gold)
    wb = load_workbook(path)
    ws = wb["Label these"]
    ws["I2"] = "N"          # row 1 labelled
    ws["I5"] = "P"          # row 4 labelled -> rows 2 and 3 are verdicts
    filled = tmp_path / "partly.xlsx"
    wb.save(filled)

    mod.read_back(filled, 1, gold_path=gold)
    out = pd.read_csv(gold)
    checked = out["checked"].astype(str).str.strip().str.lower() == "x"
    assert int(checked.sum()) == 4, "blanks between labels must count as verdicts"
    assert not checked.iloc[4:].any(), "rows past the last label must stay open"


def test_tests_never_write_to_the_real_gold_set():
    """A tripwire on the tests themselves.

    Annotation is hours of somebody's life and is not reproducible. No test
    may name the real gold set as an import target.
    """
    src = Path(__file__).read_text(encoding="utf-8")
    body = src.split("def test_tests_never_write")[0]
    for call in re.findall(r"read_back\([^)]*\)", body):
        assert "gold_path=" in call,             "read_back without gold_path -- that writes to the real gold set: " + call
